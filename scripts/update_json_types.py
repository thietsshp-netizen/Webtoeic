import openpyxl
import json
import os

file_path = '/Users/thietphamvan/hoctoeic/Webtoeic/Part 5/Part 5-tong hop/Part 5 - tong hop_phanLoai.xlsx'

def main():
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print("Loading workbook with openpyxl...")
    # data_only=False helps preserve formulas if any, but since we are modifying values, 
    # we should be careful. Usually default is fine.
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active # Assumes first sheet is active

    # Get header row
    header = [cell.value for cell in ws[1]]
    
    try:
        ai_json_idx = header.index('AI_JSON') + 1
        q_type_idx = header.index('Question_Type') + 1
    except ValueError as e:
        print(f"Error: Required column not found. Available columns: {header}")
        return

    print(f"Updating rows in sheet: {ws.title}")
    
    update_count = 0
    # ws.max_row can sometimes be larger than actual data rows if there's stray formatting
    for row_idx in range(2, ws.max_row + 1):
        json_cell = ws.cell(row=row_idx, column=ai_json_idx)
        q_type_cell = ws.cell(row=row_idx, column=q_type_idx)
        
        json_val = json_cell.value
        q_type_val = q_type_cell.value
        
        if json_val and isinstance(json_val, str) and json_val.strip().startswith('{'):
            try:
                data = json.loads(json_val)
                q_type = str(q_type_val).strip() if q_type_val else ""
                
                # Rebuild dict to insert Question_Type after questionText
                new_data = {}
                added = False
                for key, value in data.items():
                    new_data[key] = value
                    if key == 'questionText':
                        new_data['Question_Type'] = q_type
                        added = True
                
                if not added:
                    new_data['Question_Type'] = q_type
                
                # Save back to cell
                json_cell.value = json.dumps(new_data, ensure_ascii=False, indent=2)
                update_count += 1
            except Exception as e:
                print(f"Error at row {row_idx}: {e}")

    print(f"Saving workbook (Updated {update_count} rows)...")
    wb.save(file_path)
    print("Done!")

if __name__ == "__main__":
    main()
